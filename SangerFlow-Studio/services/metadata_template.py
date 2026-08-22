"""Template writer for the existing sample-metadata import contract."""

from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

from core.sequence_dataset import SequenceDataset


METADATA_TEMPLATE_HEADERS = (
    "Sample_ID",
    "Species",
    "Location",
    "Population",
    "Country",
    "Latitude",
    "Longitude",
    "Collection_Date",
    "Voucher_ID",
    "Sex",
    "Life_Stage",
    "Collector",
    "Tissue",
    "Primer",
    "Notes",
)


def write_metadata_excel_template(dataset: SequenceDataset, filepath: str | Path) -> Path:
    """Write an XLSX template in the immutable Dataset's record order."""

    if not isinstance(dataset, SequenceDataset):
        raise ValueError("metadata templates require a SequenceDataset")
    output = Path(filepath)
    if output.suffix.lower() != ".xlsx":
        output = output.with_suffix(".xlsx")
    if output.exists() and output.is_dir():
        raise ValueError(f"metadata template path is a directory: {output}")
    try:
        output.parent.mkdir(parents=True, exist_ok=True)
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = "Sample Metadata"
        headers = METADATA_TEMPLATE_HEADERS
        worksheet.append(headers)
        for record in dataset.records:
            worksheet.append((record.sequence_id,) + (None,) * (len(headers) - 1))
        for cell in worksheet[1]:
            cell.font = Font(bold=True)
        worksheet.freeze_panes = "A2"
        worksheet.auto_filter.ref = worksheet.dimensions
        for index, header in enumerate(headers, start=1):
            worksheet.column_dimensions[get_column_letter(index)].width = max(14, min(28, len(header) + 4))
        worksheet.column_dimensions[get_column_letter(1)].width = 28
        worksheet.column_dimensions[get_column_letter(len(headers))].width = 36
        workbook.save(output)
        workbook.close()
    except OSError as error:
        raise ValueError(f"could not write metadata template: {error}") from error
    return output


def write_project_metadata_excel_template(
    datasets: tuple[SequenceDataset, ...], filepath: str | Path
) -> Path:
    """Write a Project-scope template, adding Source_Batch only when needed."""

    datasets = tuple(datasets)
    if not datasets or any(not isinstance(dataset, SequenceDataset) for dataset in datasets):
        raise ValueError("metadata templates require at least one SequenceDataset")
    records = tuple(record for dataset in datasets for record in dataset.records)
    duplicate_ids = len({record.sequence_id for record in records}) != len(records)
    headers = (
        ("Sample_ID", "Source_Batch") + METADATA_TEMPLATE_HEADERS[1:]
        if duplicate_ids else METADATA_TEMPLATE_HEADERS
    )
    output = Path(filepath)
    if output.suffix.lower() != ".xlsx":
        output = output.with_suffix(".xlsx")
    try:
        output.parent.mkdir(parents=True, exist_ok=True)
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = "Sample Metadata"
        worksheet.append(headers)
        for record in records:
            worksheet.append(
                (record.sequence_id, record.metadata.get("source_batch", ""))
                + (None,) * (len(headers) - 2)
                if duplicate_ids else (record.sequence_id,) + (None,) * (len(headers) - 1)
            )
        for cell in worksheet[1]:
            cell.font = Font(bold=True)
        worksheet.freeze_panes = "A2"
        worksheet.auto_filter.ref = worksheet.dimensions
        for index, header in enumerate(headers, start=1):
            worksheet.column_dimensions[get_column_letter(index)].width = max(14, min(28, len(header) + 4))
        worksheet.column_dimensions[get_column_letter(1)].width = 28
        worksheet.column_dimensions[get_column_letter(len(headers))].width = 36
        workbook.save(output)
        workbook.close()
    except OSError as error:
        raise ValueError(f"could not write metadata template: {error}") from error
    return output
