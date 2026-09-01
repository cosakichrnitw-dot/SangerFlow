"""Studio BLAST/BOLD identification workflow integration tests."""

from __future__ import annotations

import os
import sys
import time
from pathlib import Path
from tempfile import TemporaryDirectory
from unittest.mock import patch
import unittest

from Bio.Seq import Seq
from Bio.SeqRecord import SeqRecord

os.environ.setdefault("QT_QPA_PLATFORM", "offscreen")
studio_root = Path(__file__).resolve().parents[1]
repository_root = studio_root.parent
sys.path.insert(0, str(studio_root))
sys.path.insert(0, str(repository_root))

from app.qt_runtime import configure_qt_plugins

configure_qt_plugins()

from app.app_state import AppState
from controllers.project_controller import ProjectController, _alignment_dataset_from_alignment
from core.alignment_dataset import AlignmentDataset, AlignmentRecord
from core.analysis_result import AnalysisResultType
from core.blast_filter import BlastResultFilter, apply_blast_filter
from core.blast_result import BlastAnalysisMode, BlastHit, BlastResultDataset
from core.bold_filter import BoldResultFilter, apply_bold_filter
from core.project import Project
from core.lineage import LineageSourceKind, RecordRef
from core.sequence_dataset import SequenceDataset, SequenceRecord, SourceType
from persistence.project_bundle import load_project_bundle
from PySide6.QtCore import QThread
from PySide6.QtWidgets import QApplication, QToolBar
from openpyxl import load_workbook
from views.project_view import ProjectView
from widgets.identification_service_dialogs import (
    BlastMetadataDialog, BlastMetadataSettings, BlastSettingsDialog, BlastWebsiteDialog,
    IdentificationProgressDialog, website_blast_fasta,
)
from widgets.viewers.alignment_viewer import AlignmentViewer
from widgets.viewers.dataset_viewer import DatasetViewer
from widgets.viewers.identification_result_viewers import (
    BlastResultStudioViewer,
    BoldResultStudioViewer,
)
from workflow.ncbi_blast_service import NcbiBlastProgress, NcbiBlastSettings


class IdentificationWorkflowTests(unittest.TestCase):
    @classmethod
    def setUpClass(cls) -> None:
        cls.application = QApplication.instance() or QApplication([])

    def test_sequence_dataset_runs_blast_filters_selection_creates_dataset_and_reload(self) -> None:
        dataset = SequenceDataset.from_sequence_pairs(
            "reviewed-consensus",
            "Reviewed Consensus",
            SourceType.REVIEWED_CONSENSUS,
            (("IK345", "ATGC"), ("IK346", "ATGT"), ("IK347", "TTTT")),
        )
        state, controller, view = _studio(Project.create("project", "Project").add_dataset(dataset))
        controller.configure_identification_runners(blast_runner=_blast_runner)

        result = controller.run_blast_for_dataset(dataset)
        self.application.processEvents()

        self.assertEqual(result.parent_dataset_id, "reviewed-consensus")
        self.assertTrue(state.current_project.has_analysis_result(result.result_id))
        self.assertIsInstance(state.active_viewer, BlastResultStudioViewer)

        viewer = state.active_viewer
        self.assertIn("BLAST completed. Stored in Project Results", viewer._project_storage_feedback.text())
        self.assertIn("Save Project to persist it to disk.", viewer._project_storage_feedback.text())
        viewer._scientific_name.setText("Rhynchobatus springeri")
        viewer._min_identity.setText("98")
        viewer._min_coverage.setText("95")
        selection = viewer.apply_filter_from_fields()

        self.assertEqual(selection.selected_query_ids, ("IK345", "IK346"))
        subset = controller.create_dataset_from_blast_result_selection(
            viewer,
            selection,
            name="Rhynchobatus springeri",
        )

        self.assertEqual(subset.sequence_ids, ("IK345", "IK346"))
        self.assertEqual(subset.metadata["parent_dataset_id"], "reviewed-consensus")
        self.assertEqual(subset.metadata["derivation_type"], "BLAST_SELECTION")
        self.assertEqual(subset.metadata["source_analysis"], "BLAST")
        self.assertEqual(subset.metadata["selected_record_ids"], ("IK345", "IK346"))
        self.assertTrue(state.current_project.has_dataset(subset.dataset_id))
        self.assertEqual(
            tuple(
                relation.source_kind
                for relation in state.current_project.get_entry(subset.dataset_id).lineage_relations
            ),
            (LineageSourceKind.DATASET, LineageSourceKind.ANALYSIS_RESULT),
        )

        with TemporaryDirectory() as directory:
            bundle = Path(directory) / "project.sangerflow"
            controller.save_project_bundle(str(bundle))
            loaded = load_project_bundle(bundle)
            try:
                self.assertTrue(loaded.project.has_analysis_result(result.result_id))
                self.assertTrue(loaded.repository.has_result(result.result_id))
                self.assertTrue(loaded.project.has_dataset(subset.dataset_id))
                self.assertEqual(
                    loaded.project.get_dataset(subset.dataset_id).metadata["source_result_id"],
                    result.result_id,
                )
            finally:
                loaded.cleanup()
        view.close()

    def test_blast_xml_import_feedback_identifies_project_results_storage(self) -> None:
        dataset = SequenceDataset.from_sequence_pairs(
            "reviewed-consensus", "Reviewed Consensus", SourceType.REVIEWED_CONSENSUS,
            (("IK345", "ATGC"),),
        )
        state, controller, view = _studio(Project.create("project", "Project").add_dataset(dataset))
        imported = BlastResultDataset(
            result_id="web-xml-result",
            name="Imported Web BLAST",
            parent_dataset_id=dataset.dataset_id,
            analysis_mode=BlastAnalysisMode.IDENTIFICATION,
            database="nt",
            hits=(BlastHit("IK345", "ACC-1", "Rhynchobatus springeri", "Rhynchobatus springeri", 99.0, 100.0, 1e-20, 4, "nt"),),
        )
        with patch(
            "controllers.project_controller.import_ncbi_blast_xml",
            return_value=(imported, object()),
        ):
            controller.import_ncbi_blast_xml_for_dataset(dataset, "/tmp/example.xml")

        self.assertTrue(state.current_project.has_analysis_result("web-xml-result"))
        self.assertIsInstance(state.active_viewer, BlastResultStudioViewer)
        self.assertIn("NCBI BLAST XML imported. Stored in Project Results", state.active_viewer._project_storage_feedback.text())
        self.assertIn("Save Project to persist it to disk.", state.active_viewer._project_storage_feedback.text())
        view.close()

    def test_alignment_dataset_runs_bold_with_ungapped_query_and_creates_selection_dataset(self) -> None:
        parent = SequenceDataset.from_sequence_pairs(
            "imported-fasta",
            "Imported FASTA",
            SourceType.IMPORTED_FASTA,
            (("IK345", "ATGCATCG"), ("IK346", "ATGTATCG")),
        )
        alignment = AlignmentDataset(
            alignment_id="alignment",
            name="Alignment",
            parent_dataset_id=parent.dataset_id,
            records=(
                AlignmentRecord("IK345", "IK345", "ATGC--ATCG"),
                AlignmentRecord("IK346", "IK346", "ATGT--ATCG"),
            ),
        )
        project = (
            Project.create("project", "Project")
            .add_dataset(parent)
            .add_dataset(alignment, parent_dataset_id=parent.dataset_id)
        )
        calls: list[str] = []

        def bold_runner(sequence: str) -> dict[str, object]:
            calls.append(sequence)
            query_id = "IK345" if sequence == "ATGCATCG" else "IK346"
            return {
                "query_id": query_id,
                "process_id": f"BOLD:{query_id}",
                "species_name": "Rhynchobatus springeri",
                "genus": "Rhynchobatus",
                "similarity": 99.2,
                "bin_uri": "BOLD:AAA001",
            }

        state, controller, view = _studio(project)
        controller.configure_identification_runners(bold_runner=bold_runner)

        result = controller.run_bold_for_dataset(alignment)
        self.application.processEvents()

        self.assertEqual(calls, ["ATGCATCG", "ATGTATCG"])
        self.assertEqual(result.parent_dataset_id, "alignment")
        self.assertEqual(alignment.get_record("IK345").aligned_sequence, "ATGC--ATCG")
        self.assertIsInstance(state.active_viewer, BoldResultStudioViewer)

        viewer = state.active_viewer
        viewer._species_name.setText("Rhynchobatus springeri")
        viewer._min_similarity.setText("99")
        selection = viewer.apply_filter_from_fields()
        subset = controller.create_dataset_from_bold_result_selection(
            viewer,
            selection,
            name="BOLD springeri subset",
        )

        self.assertEqual(subset.sequence_ids, ("IK345", "IK346"))
        self.assertEqual(tuple(record.sequence for record in subset.records), ("ATGCATCG", "ATGTATCG"))
        self.assertEqual(subset.metadata["derivation_type"], "BOLD_SELECTION")
        self.assertEqual(subset.metadata["source_analysis"], "BOLD")
        self.assertEqual(state.current_project.lineage(subset.dataset_id), ("imported-fasta", "alignment", subset.dataset_id))
        self.assertEqual(
            tuple(
                relation.source_kind
                for relation in state.current_project.get_entry(subset.dataset_id).lineage_relations
            ),
            (LineageSourceKind.DATASET, LineageSourceKind.ANALYSIS_RESULT),
        )
        view.close()

    def test_dataset_and_alignment_actions_expose_blast_but_not_bold_online(self) -> None:
        dataset = SequenceDataset.from_sequence_pairs(
            "imported",
            "Imported",
            SourceType.IMPORTED_FASTA,
            (("IK345", "ATGC"),),
        )
        alignment = AlignmentDataset(
            alignment_id="alignment",
            name="Alignment",
            parent_dataset_id=dataset.dataset_id,
            records=(AlignmentRecord("IK345", "IK345", "AT-GC"),),
        )
        project = Project.create("project", "Project").add_dataset(dataset).add_dataset(
            alignment,
            parent_dataset_id=dataset.dataset_id,
        )
        state, controller, view = _studio(project)
        toolbar = QToolBar()
        view.action_manager.attach_toolbar(toolbar)

        dataset_viewer = DatasetViewer(dataset, view.viewer_context)
        state.set_active_viewer(dataset_viewer)
        self.application.processEvents()

        self.assertIsNotNone(view.action_manager.action("dataset.run_blast"))
        self.assertIsNone(view.action_manager.action("dataset.run_bold"))

        alignment_viewer = AlignmentViewer(alignment, context=view.viewer_context)
        state.set_active_viewer(alignment_viewer)
        self.application.processEvents()

        self.assertIsNotNone(view.action_manager.action("alignment.run_blast"))
        self.assertIsNone(view.action_manager.action("alignment.run_bold"))
        view.close()

    def test_dataset_viewer_run_blast_action_uses_interactive_controller_path(self) -> None:
        dataset = SequenceDataset.from_sequence_pairs(
            "imported",
            "Imported",
            SourceType.IMPORTED_FASTA,
            (("IK345", "ATGC"),),
        )
        state, controller, view = _studio(Project.create("project", "Project").add_dataset(dataset))
        calls: list[object] = []

        def interactive(selected_dataset: object, *, parent_widget: object | None = None) -> None:
            calls.append((selected_dataset, parent_widget))

        controller.run_blast_for_dataset_interactive = interactive  # type: ignore[method-assign]
        toolbar = QToolBar()
        view.action_manager.attach_toolbar(toolbar)
        dataset_viewer = DatasetViewer(dataset, view.viewer_context)
        state.set_active_viewer(dataset_viewer)
        self.application.processEvents()

        action = view.action_manager.action("dataset.run_blast")
        self.assertIsNotNone(action)
        action.trigger()

        self.assertEqual(calls[0][0], dataset)
        self.assertIs(calls[0][1], dataset_viewer)
        view.close()

    def test_dataset_viewer_passes_include_checkboxes_to_blast_controller(self) -> None:
        dataset = SequenceDataset.from_sequence_pairs(
            "imported", "Imported", SourceType.IMPORTED_FASTA,
            (("IK345", "ATGC"), ("IK346", "ATGT")),
        )
        state, controller, view = _studio(Project.create("project", "Project").add_dataset(dataset))
        calls: list[object] = []

        def interactive(selected_dataset: object, *, included_record_ids, parent_widget: object | None = None) -> None:
            calls.append((selected_dataset, included_record_ids, parent_widget))

        controller.run_blast_for_dataset_interactive = interactive  # type: ignore[method-assign]
        toolbar = QToolBar()
        view.action_manager.attach_toolbar(toolbar)
        dataset_viewer = DatasetViewer(dataset, view.viewer_context)
        dataset_viewer._included_record_ids = {"IK346"}
        state.set_active_viewer(dataset_viewer)
        self.application.processEvents()

        view.action_manager.action("dataset.run_blast").trigger()

        self.assertEqual(calls[0][0], dataset)
        self.assertEqual(calls[0][1], frozenset({"IK346"}))
        self.assertIs(calls[0][2], dataset_viewer)
        view.close()

    def test_blast_settings_and_progress_dialogs_are_gui_safe(self) -> None:
        settings_dialog = BlastSettingsDialog(query_count=119)
        settings = settings_dialog.settings()

        self.assertEqual(settings.program, "blastn")
        self.assertEqual(settings.database, "nt")
        self.assertEqual(settings.max_target_sequences, 10)

        progress_dialog = IdentificationProgressDialog()
        progress_dialog.update_progress(
            NcbiBlastProgress(
                state="Retrieving",
                query_id="IK345",
                completed=3,
                total=10,
                successful=2,
                no_hit=1,
                failed=0,
                rid="RID123",
            )
        )
        self.assertFalse(progress_dialog.cancelled)
        progress_dialog.cancel()
        self.assertTrue(progress_dialog.cancelled)
        settings_dialog.close()
        progress_dialog.close()

    def test_blast_dialog_defaults_to_included_records_and_validates_custom_database(self) -> None:
        dialog = BlastSettingsDialog(query_count=10, included_query_count=3)
        self.assertEqual(dialog.query_scope, "included")
        self.assertEqual(dialog.final_query_count, 3)
        self.assertIn("Final query count: 3", dialog._summary.text())
        dialog._query_scope.setCurrentIndex(1)
        self.assertEqual(dialog.query_scope, "all")
        self.assertEqual(dialog.final_query_count, 10)
        dialog._database.setCurrentText("Custom…")
        with self.assertRaisesRegex(ValueError, "database"):
            dialog.settings()
        dialog._custom_database.setText("nt_custom")
        self.assertEqual(dialog.settings().database, "nt_custom")
        dialog.close()

    def test_website_fasta_preserves_exact_ids_and_record_order(self) -> None:
        dataset = SequenceDataset.from_sequence_pairs(
            "input", "Input", SourceType.IMPORTED_FASTA,
            (("C2_FishF1", "ATGC"), ("C3 FishF1", "TTAA")),
        )
        fasta = website_blast_fasta(dataset)
        self.assertEqual(fasta, ">C2_FishF1\nATGC\n>C3 FishF1\nTTAA\n")
        dialog = BlastWebsiteDialog(dataset, on_import_xml=lambda _path: None)
        self.assertEqual(dialog.copy_fasta_to_clipboard(), fasta)
        self.assertEqual(self.application.clipboard().text(), fasta)
        dialog.close()

    def test_blast_included_records_are_the_only_queries_submitted(self) -> None:
        dataset = SequenceDataset.from_sequence_pairs(
            "input", "Input", SourceType.IMPORTED_FASTA,
            (("IK345", "ATGC"), ("IK346", "TTTT"), ("IK347", "ATGA")),
        )
        state, controller, view = _studio(Project.create("project", "Project").add_dataset(dataset))
        submitted: list[str] = []

        def runner(sequence: str) -> list[dict[str, object]]:
            submitted.append(sequence)
            return _blast_runner(sequence)

        controller.configure_identification_runners(blast_runner=runner)
        result = controller.run_blast_for_dataset(dataset, included_record_ids={"IK345", "IK347"})

        self.assertEqual(submitted, ["ATGC", "ATGA"])
        self.assertEqual(result.query_ids(), ("IK345", "IK347"))
        self.assertEqual(result.parent_dataset_id, dataset.dataset_id)
        self.assertEqual(result.metadata["selected_record_ids"], ("IK345", "IK347"))
        view.close()

    def test_blast_metadata_dialog_requires_fields_and_preserves_user_selection(self) -> None:
        dialog = BlastMetadataDialog(dataset_name="Reviewed", query_count=2)
        dialog._field_boxes["blast_evalue"].setChecked(False)
        settings = dialog.settings()
        self.assertNotIn("blast_evalue", settings.fields)
        self.assertIn("blast_accession", settings.fields)
        for checkbox in dialog._field_boxes.values():
            checkbox.setChecked(False)
        with self.assertRaisesRegex(ValueError, "at least one"):
            dialog.settings()
        dialog.close()

    def test_async_blast_worker_queues_result_tab_creation_to_main_gui_thread(self) -> None:
        dataset = SequenceDataset.from_sequence_pairs(
            "async-input",
            "Async input",
            SourceType.IMPORTED_FASTA,
            (("IK345", "ATGC"),),
        )
        state, controller, view = _studio(Project.create("project", "Project").add_dataset(dataset))
        progress_dialog = IdentificationProgressDialog()
        worker_threads: list[object] = []
        completion_threads: list[object] = []
        original_complete = controller._complete_async_blast

        def record_complete(*args, **kwargs):
            completion_threads.append(QThread.currentThread())
            return original_complete(*args, **kwargs)

        controller._complete_async_blast = record_complete  # type: ignore[method-assign]

        class FakeRunner:
            def __init__(self, _settings) -> None:
                pass

            def run_dataset(self, query_dataset, *, analysis_mode, progress, should_cancel):
                worker_threads.append(QThread.currentThread())
                progress(NcbiBlastProgress(state="Working", completed=0, total=1))
                if should_cancel():
                    raise AssertionError("test worker was unexpectedly cancelled")
                return BlastResultDataset(
                    result_id="async-blast-result",
                    name="Async BLAST",
                    parent_dataset_id=query_dataset.dataset_id,
                    hits=(
                        BlastHit(
                            query_id="IK345",
                            hit_accession="ACC-1",
                            scientific_name="Rhynchobatus springeri",
                            organism="Rhynchobatus springeri",
                            identity=99.0,
                            query_coverage=100.0,
                            evalue=0.0,
                            alignment_length=4,
                            database="nt",
                        ),
                    ),
                    analysis_mode=BlastAnalysisMode.IDENTIFICATION,
                    database="nt",
                )

        with patch("controllers.identification_workers.NcbiBlastRunner", FakeRunner), patch(
            "controllers.project_controller.QMessageBox.critical"
        ) as critical:
            controller._start_async_blast_worker(
                dataset,
                NcbiBlastSettings(),
                source_dataset=dataset,
                parent_dataset_id=dataset.dataset_id,
                progress_dialog=progress_dialog,
            )
            deadline = time.monotonic() + 3.0
            while not state.current_project.has_analysis_result("async-blast-result"):
                self.application.processEvents()
                if time.monotonic() >= deadline:
                    if critical.called:
                        self.fail(f"async BLAST registration failed: {critical.call_args.args[2]}")
                    self.fail("timed out waiting for the queued BLAST result")
                time.sleep(0.01)
            self.application.processEvents()

        self.assertEqual(len(worker_threads), 1)
        self.assertNotEqual(worker_threads[0], self.application.thread())
        self.assertEqual(completion_threads, [self.application.thread()])
        self.assertIsInstance(state.active_viewer, BlastResultStudioViewer)
        self.assertGreaterEqual(view.tab_manager._tabs.count(), 3)
        explorer = view.widget(0)
        self.assertEqual(explorer.topLevelItem(0).child(2).childCount(), 1)
        self.assertEqual(explorer.topLevelItem(0).child(2).child(0).text(0), "Async BLAST")
        view.close()

    def test_project_explorer_reopens_persisted_blast_result_payload(self) -> None:
        dataset = SequenceDataset.from_sequence_pairs(
            "fas-dataset",
            "FAS Dataset",
            SourceType.IMPORTED_FASTA,
            (("IK345", "ATGC"),),
        )
        state, controller, view = _studio(Project.create("project", "Project").add_dataset(dataset))
        controller.configure_identification_runners(blast_runner=_blast_runner)
        result = controller.run_blast_for_dataset(dataset)
        self.application.processEvents()

        explorer = view.widget(0)
        result_item = explorer.topLevelItem(0).child(2).child(0)
        explorer.setCurrentItem(result_item)
        self.application.processEvents()

        self.assertIsInstance(state.active_viewer, BlastResultStudioViewer)
        self.assertEqual(state.active_viewer.result.result_id, result.result_id)
        view.close()

    def test_reloaded_blast_result_viewer_exports_excel_and_tsv(self) -> None:
        dataset = SequenceDataset.from_sequence_pairs(
            "fas-dataset",
            "FAS Dataset",
            SourceType.IMPORTED_FASTA,
            (("IK345", "ATGC"),),
        )
        state, controller, view = _studio(Project.create("project", "Project").add_dataset(dataset))
        controller.configure_identification_runners(blast_runner=_blast_runner)
        result = controller.run_blast_for_dataset(dataset)
        self.application.processEvents()

        with TemporaryDirectory() as directory:
            bundle = Path(directory) / "project.sangerflow"
            controller.save_project_bundle(str(bundle))
            loaded = load_project_bundle(bundle)
            try:
                loaded_state, loaded_controller, loaded_view = _studio(loaded.project)
                loaded_state.set_repository(loaded.repository)
                explorer = loaded_view.widget(0)
                result_item = explorer.topLevelItem(0).child(2).child(0)
                explorer.setCurrentItem(result_item)
                self.application.processEvents()
                viewer = loaded_state.active_viewer

                excel_target = Path(directory) / "blast_result"
                tsv_target = Path(directory) / "blast_result"
                with patch(
                    "widgets.viewers.identification_result_viewers.QFileDialog.getSaveFileName",
                    return_value=(str(excel_target), "Excel Workbook (*.xlsx)"),
                ):
                    excel_path = viewer.request_export_excel()
                with patch(
                    "widgets.viewers.identification_result_viewers.QFileDialog.getSaveFileName",
                    return_value=(str(tsv_target), "TSV files (*.tsv)"),
                ):
                    tsv_path = viewer.request_export_tsv()

                workbook = load_workbook(excel_path)
                self.assertEqual(workbook.sheetnames, ["Summary", "Query Summary", "All Hits"])
                self.assertIn("Rhynchobatus springeri", Path(tsv_path).read_text(encoding="utf-8"))
                self.assertEqual(viewer.result.result_id, result.result_id)
                loaded_view.close()
            finally:
                loaded.cleanup()
        view.close()

    def test_bold_result_viewer_exports_excel_and_tsv(self) -> None:
        dataset = SequenceDataset.from_sequence_pairs(
            "fas-dataset",
            "FAS Dataset",
            SourceType.IMPORTED_FASTA,
            (("IK345", "ATGC"),),
        )
        state, controller, view = _studio(Project.create("project", "Project").add_dataset(dataset))
        controller.configure_identification_runners(
            bold_runner=lambda _sequence: {
                "query_id": "IK345",
                "process_id": "BOLD:IK345",
                "species_name": "Rhynchobatus springeri",
                "genus": "Rhynchobatus",
                "similarity": 99.2,
                "bin_uri": "BOLD:AAA001",
            }
        )
        controller.run_bold_for_dataset(dataset)
        self.application.processEvents()
        viewer = state.active_viewer

        with TemporaryDirectory() as directory:
            excel_target = Path(directory) / "bold_result"
            tsv_target = Path(directory) / "bold_result"
            with patch(
                "widgets.viewers.identification_result_viewers.QFileDialog.getSaveFileName",
                return_value=(str(excel_target), "Excel Workbook (*.xlsx)"),
            ):
                excel_path = viewer.request_export_excel()
            with patch(
                "widgets.viewers.identification_result_viewers.QFileDialog.getSaveFileName",
                return_value=(str(tsv_target), "TSV files (*.tsv)"),
            ):
                tsv_path = viewer.request_export_tsv()

            workbook = load_workbook(excel_path)
            self.assertEqual(
                workbook.sheetnames,
                ["Summary", "Taxonomic Assignment", "Reference Information", "All Hits"],
            )
            self.assertIn("Rhynchobatus springeri", Path(tsv_path).read_text(encoding="utf-8"))
        view.close()

    def test_blast_export_error_is_reported_with_message_box(self) -> None:
        dataset = SequenceDataset.from_sequence_pairs(
            "fas-dataset",
            "FAS Dataset",
            SourceType.IMPORTED_FASTA,
            (("IK345", "ATGC"),),
        )
        state, controller, view = _studio(Project.create("project", "Project").add_dataset(dataset))
        controller.configure_identification_runners(blast_runner=_blast_runner)
        controller.run_blast_for_dataset(dataset)
        self.application.processEvents()
        viewer = state.active_viewer

        with TemporaryDirectory() as directory:
            with patch(
                "widgets.viewers.identification_result_viewers.QFileDialog.getSaveFileName",
                return_value=(str(Path(directory) / "blast.xlsx"), "Excel Workbook (*.xlsx)"),
            ), patch(
                "widgets.viewers.identification_result_viewers.export_blast_result_to_excel",
                side_effect=ValueError("cannot write export"),
            ), patch("widgets.viewers.identification_result_viewers.QMessageBox.warning") as warning:
                result = viewer.request_export_excel()

        self.assertIsNone(result)
        warning.assert_called_once()
        self.assertIn("cannot write export", warning.call_args.args[2])
        view.close()

    def test_core_filters_remain_reused_by_studio_viewers(self) -> None:
        dataset = SequenceDataset.from_sequence_pairs(
            "fas-dataset",
            "FAS Dataset",
            SourceType.IMPORTED_FASTA,
            (("IK345", "ATGC"), ("IK346", "ATGT")),
        )
        state, controller, view = _studio(Project.create("project", "Project").add_dataset(dataset))
        controller.configure_identification_runners(blast_runner=_blast_runner)
        result = controller.run_blast_for_dataset(dataset)

        selection = apply_blast_filter(
            result,
            BlastResultFilter(scientific_name="Rhynchobatus springeri", min_identity=98),
        )

        self.assertEqual(selection.selected_query_ids, ("IK345", "IK346"))

        controller.configure_identification_runners(bold_runner=lambda sequence: {
            "query_id": "IK345" if sequence == "ATGC" else "IK346",
            "process_id": f"BOLD:{sequence}",
            "species_name": "Rhynchobatus springeri",
            "similarity": 99.0,
        })
        bold_result = controller.run_bold_for_dataset(dataset)
        bold_selection = apply_bold_filter(
            bold_result,
            BoldResultFilter(species_name="Rhynchobatus springeri", min_similarity=98),
        )

        self.assertEqual(bold_selection.selected_query_ids, ("IK345", "IK346"))
        view.close()

    def test_blast_metadata_creates_current_metadata_revision_without_overwriting_species(self) -> None:
        dataset = SequenceDataset.from_sequence_pairs(
            "reviewed", "Reviewed", SourceType.REVIEWED_CONSENSUS,
            (("IK345", "ATGC"), ("IK346", "TTTT")),
        )
        state, controller, view = _studio(Project.create("project", "Project").add_dataset(dataset))
        controller.configure_identification_runners(blast_runner=_blast_runner)
        result = controller.run_blast_for_dataset(dataset)
        with self.assertRaisesRegex(ValueError, "not present in this result"):
            controller.apply_blast_result_metadata(
                result,
                BlastMetadataSettings(),
                selected_query_ids=("unknown-query",),
            )
        derived = controller.apply_blast_result_metadata(
            result, BlastMetadataSettings(minimum_identity=98.0, minimum_coverage=90.0),
        )

        accepted = derived.get_record("IK345").metadata
        uncertain = derived.get_record("IK346").metadata
        self.assertEqual(accepted["blast_scientific_name"], "Rhynchobatus springeri")
        self.assertEqual(accepted["blast_identification_status"], "accepted")
        self.assertEqual(uncertain["blast_identification_status"], "uncertain")
        self.assertNotIn("species", accepted)
        entry = state.current_project.get_entry(derived.dataset_id)
        self.assertEqual(entry.logical_id, dataset.dataset_id)
        self.assertEqual(entry.revision_number, 2)
        self.assertFalse(state.current_project.is_current_revision(dataset.dataset_id))
        self.assertTrue(state.current_project.is_current_revision(derived.dataset_id))
        view.close()

    def test_blast_metadata_keeps_best_hit_text_separate_from_scientific_name(self) -> None:
        dataset = SequenceDataset.from_sequence_pairs(
            "reviewed", "Reviewed", SourceType.REVIEWED_CONSENSUS,
            (("IK345", "ATGC"),),
        )
        result = BlastResultDataset(
            result_id="blast-result",
            name="BLAST result",
            parent_dataset_id=dataset.dataset_id,
            analysis_mode=BlastAnalysisMode.IDENTIFICATION,
            database="nt",
            hits=(BlastHit(
                "IK345", "ACC-1", "Rhynchobatus australiae", "Rhynchobatus australiae",
                99.0, 100.0, 1e-40, 4, "nt", 200.0,
                "Rhynchobatus australiae mitochondrion, complete genome",
            ),),
        )
        state, controller, view = _studio(Project.create("project", "Project").add_dataset(dataset))
        derived = controller.apply_blast_result_metadata(result, BlastMetadataSettings())

        metadata = derived.get_record("IK345").metadata
        self.assertEqual(metadata["blast_best_hit"], "Rhynchobatus australiae mitochondrion, complete genome")
        self.assertEqual(metadata["blast_scientific_name"], "Rhynchobatus australiae")
        view.close()

    def test_selected_blast_queries_only_receive_metadata_and_survive_reload(self) -> None:
        dataset = SequenceDataset(
            "reviewed", "Reviewed", SourceType.REVIEWED_CONSENSUS,
            (
                SequenceRecord("IK345", "ATGC", metadata={"source_batch": "Cirebon"}),
                SequenceRecord("IK346", "TTTT", metadata={"source_batch": "Cirebon"}),
            ),
            metadata={"source_batch": "Cirebon"},
        )
        state, controller, view = _studio(Project.create("project", "Project").add_dataset(dataset))
        controller.configure_identification_runners(blast_runner=_blast_runner)
        result = controller.run_blast_for_dataset(dataset)
        derived = controller.apply_blast_result_metadata(
            result,
            BlastMetadataSettings(minimum_identity=98.0, minimum_coverage=90.0),
            selected_query_ids=("IK345",),
        )

        self.assertEqual(derived.get_record("IK345").metadata["blast_accession"], "ACC-ATGC")
        self.assertEqual(derived.get_record("IK345").metadata["source_batch"], "Cirebon")
        self.assertNotIn("blast_accession", derived.get_record("IK346").metadata)
        self.assertEqual(derived.metadata["blast_metadata_selected_query_ids"], ("IK345",))

        selected = controller.create_dataset_from_project_record_refs(
            (RecordRef(derived.dataset_id, "IK345"),),
            dataset_id="cirebon_selected",
            name="Cirebon selected",
        )
        self.assertEqual(selected.get_record("IK345").metadata["source_batch"], "Cirebon")
        self.assertEqual(selected.get_record("IK345").metadata["blast_accession"], "ACC-ATGC")
        alignment = _alignment_dataset_from_alignment(
            (SeqRecord(Seq("ATGC"), id="IK345"),),
            selected,
            alignment_id="cirebon_alignment",
            name="Cirebon alignment",
            metadata={"source_batch": "Cirebon"},
        )
        state.replace_project(
            state.current_project.add_dataset(alignment, parent_dataset_id=selected.dataset_id),
            dirty=True,
        )
        with TemporaryDirectory() as directory:
            bundle = Path(directory) / "selected-metadata.sangerflow"
            controller.save_project_bundle(str(bundle))
            loaded = load_project_bundle(bundle)
            try:
                restored = loaded.project.get_dataset(derived.dataset_id)
                self.assertEqual(restored.get_record("IK345").metadata["blast_accession"], "ACC-ATGC")
                self.assertEqual(restored.get_record("IK345").metadata["source_batch"], "Cirebon")
                self.assertNotIn("blast_accession", restored.get_record("IK346").metadata)
                restored_selected = loaded.project.get_dataset(selected.dataset_id)
                self.assertEqual(restored_selected.get_record("IK345").metadata["source_batch"], "Cirebon")
                restored_alignment = loaded.project.get_dataset(alignment.alignment_id)
                self.assertEqual(restored_alignment.metadata["source_batch"], "Cirebon")
                self.assertEqual(restored_alignment.get_record("IK345").metadata["source_batch"], "Cirebon")
            finally:
                loaded.cleanup()
        view.close()

    def test_blast_result_table_lists_ranked_hits_with_supported_fields(self) -> None:
        result = BlastResultDataset(
            result_id="ranked-blast",
            name="Ranked BLAST",
            parent_dataset_id="input",
            analysis_mode=BlastAnalysisMode.IDENTIFICATION,
            database="nt",
            hits=(
                BlastHit("IK345", "ACC-1", "Species one", "Species one", 99.0, 100.0, 1e-20, 4, "nt", 100.0),
                BlastHit("IK345", "ACC-2", "Species two", "Species two", 98.0, 95.0, 1e-10, 4, "nt", 80.0),
            ),
        )
        viewer = BlastResultStudioViewer(result)

        self.assertEqual(viewer._table.rowCount(), 1)
        self.assertEqual(viewer._apply_metadata_button.text(), "Apply Selected Top Hits to Metadata…")
        self.assertEqual(
            [viewer._table.horizontalHeaderItem(column).text() for column in range(viewer._table.columnCount())],
            [
                "Selected", "Query ID", "Rank", "Scientific Name", "Accession",
                "Identity %", "Coverage %", "Alignment length", "E-value", "Bit Score", "Description",
            ],
        )
        self.assertEqual(viewer._table.item(0, 2).text(), "1")
        viewer._top_hit_only.setChecked(False)
        self.assertEqual(viewer._table.rowCount(), 2)
        self.assertFalse(viewer._apply_metadata_button.isEnabled())
        self.assertEqual(viewer._table.item(1, 2).text(), "2")
        self.assertEqual(viewer._table.item(1, 7).text(), "4")
        viewer.close()

    def test_blast_filter_uses_stable_query_and_accession_identity_then_clears(self) -> None:
        result = BlastResultDataset(
            result_id="filter-identities",
            name="Filter identities",
            parent_dataset_id="input",
            analysis_mode=BlastAnalysisMode.IDENTIFICATION,
            database="nt",
            hits=(
                BlastHit("IK345", "ACC-COI-1", "Species one", "Species one", 99.0, 100.0, 1e-20, 4, "nt"),
                BlastHit("IK346", "ACC-COI-2", "Species two", "Species two", 98.0, 99.0, 1e-20, 4, "nt"),
            ),
        )
        viewer = BlastResultStudioViewer(result)
        viewer._query_text.setText("345")
        viewer._accession.setText("COI-1")
        selection = viewer.apply_filter_from_fields()

        self.assertEqual(selection.selected_query_ids, ("IK345",))
        self.assertEqual(viewer._table.rowCount(), 1)
        self.assertEqual(viewer._table.item(0, 1).text(), "IK345")

        viewer.clear_filter()
        self.assertEqual(viewer._filtered_ids, ("IK345", "IK346"))
        self.assertEqual(viewer._query_text.text(), "")
        self.assertEqual(viewer._accession.text(), "")
        viewer.close()


def _studio(project: Project) -> tuple[AppState, ProjectController, ProjectView]:
    state = AppState()
    controller = ProjectController(state)
    view = ProjectView(state, controller)
    controller.open_project(project)
    return state, controller, view


def _blast_runner(sequence: str) -> list[dict[str, object]]:
    species = "Rhynchobatus springeri" if sequence.startswith("ATG") else "Dasyatis kuhlii"
    return [
        {
            "species": species,
            "identity": 99.1 if species.startswith("Rhynchobatus") else 90.0,
            "coverage": 97.5,
            "alignment_length": len(sequence),
            "e_value": 1e-40,
            "accession": f"ACC-{sequence}",
            "bit_score": 500,
        }
    ]


if __name__ == "__main__":
    unittest.main()
