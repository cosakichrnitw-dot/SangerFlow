"""Focused P0 Studio Project UX integration checks."""

from __future__ import annotations

import os
import sys
from pathlib import Path
from tempfile import TemporaryDirectory
import unittest

os.environ.setdefault("QT_QPA_PLATFORM", "offscreen")
studio_root = Path(__file__).resolve().parents[1]
repository_root = studio_root.parent
sys.path.insert(0, str(studio_root))
sys.path.insert(0, str(repository_root))

from app.qt_runtime import configure_qt_plugins
configure_qt_plugins()
from PySide6.QtWidgets import QApplication

from app.app_state import AppState
from controllers.project_controller import ProjectController
from core.analysis_result import AnalysisResult, AnalysisResultType
from core.project import Project
from core.sequence_dataset import SequenceDataset, SequenceRecord, SourceType
from services.metadata_template import METADATA_TEMPLATE_HEADERS
from services.metadata_template import write_project_metadata_excel_template
from views.project_view import ProjectView


def _dataset(identifier: str, name: str) -> SequenceDataset:
    return SequenceDataset.from_sequence_pairs(
        identifier,
        name,
        SourceType.IMPORTED_FASTA,
        (("sample-b", "ATGC"), ("sample-a", "ATGT")),
    )


class ProjectUxP0Tests(unittest.TestCase):
    def setUp(self) -> None:
        self.application = QApplication.instance() or QApplication([])
        self.state = AppState()
        self.controller = ProjectController(self.state)
        self.view = ProjectView(self.state, self.controller)

    def tearDown(self) -> None:
        self.state.close_project()

    def test_new_workspace_creates_bundle_and_standard_directories(self) -> None:
        with TemporaryDirectory() as directory:
            project = self.controller.create_project("Wedgefish 2026", location=directory)
            workspace = Path(directory) / "Wedgefish_2026"
            self.assertEqual(project.name, "Wedgefish 2026")
            self.assertTrue((workspace / "Wedgefish_2026.sangerflow").is_file())
            for name in ("Raw_Data", "Exports", "Metadata", "Reports"):
                self.assertTrue((workspace / name).is_dir())
            self.assertEqual(self.controller.export_default_directory(), str(workspace / "Exports"))
            self.assertEqual(self.controller.metadata_default_directory(), str(workspace / "Metadata"))

    def test_ab1_copy_mode_uses_raw_data_and_avoids_filename_collision(self) -> None:
        with TemporaryDirectory() as directory:
            source = Path(directory) / "source"
            source.mkdir()
            ab1 = source / "sample.ab1"
            ab1.write_bytes(b"AB1 placeholder")
            self.controller.create_project("Project", location=directory)
            first = self.controller._resolve_ab1_source_files((ab1,), "copy")
            second = self.controller._resolve_ab1_source_files((ab1,), "copy")
            self.assertEqual(first[0].name, "sample.ab1")
            self.assertEqual(second[0].name, "sample_2.ab1")
            self.assertEqual(first[0].read_bytes(), b"AB1 placeholder")

    def test_close_project_resets_project_and_viewer_state(self) -> None:
        self.state.set_project(Project.create("p", "Project"), dirty=True)
        self.controller.close_project()
        self.assertIsNone(self.state.current_project)
        self.assertIsNone(self.state.current_bundle_path)
        self.assertFalse(self.state.is_dirty)
        self.assertEqual(self.state.active_tab, "Welcome")

    def test_explorer_builds_working_sections_and_search_keeps_matching_section_visible(self) -> None:
        root = _dataset("raw", "Raw AB1 Run")
        child = _dataset("reviewed", "Reviewed Consensus")
        project = Project.create("p", "Project").add_dataset(root).add_dataset(
            child, parent_dataset_id=root.dataset_id
        )
        self.state.set_project(project)
        self.application.processEvents()
        explorer = self.view.widget(0)
        project_root = explorer.topLevelItem(0)
        working = project_root.child(0)
        self.assertEqual(working.text(0), "Working Datasets")
        self.assertEqual(tuple(working.child(index).text(0) for index in range(working.childCount())), ("Raw AB1 Run", "Reviewed Consensus"))
        explorer._search.setText("reviewed")
        self.assertFalse(working.isHidden())
        self.assertFalse(working.child(1).isHidden())

    def test_graph_uses_formal_dataset_and_result_edges(self) -> None:
        dataset = _dataset("source", "Source")
        project = Project.create("p", "Project").add_dataset(dataset).add_analysis_result(
            AnalysisResult("blast", "BLAST", AnalysisResultType.BLAST, dataset.dataset_id)
        )
        self.state.set_project(project)
        self.application.processEvents()
        graph = self.view.widget(1)._project_summary
        self.assertGreaterEqual(len(graph.scene().items()), 3)

    def test_dataset_rename_remove_and_metadata_template_are_project_operations(self) -> None:
        with TemporaryDirectory() as directory:
            project = self.controller.create_project("Project", location=directory)
            dataset = _dataset("source", "Original")
            self.state.replace_project(project.add_dataset(dataset))
            renamed = self.controller.rename_dataset(dataset.dataset_id, "Display only")
            self.assertEqual(renamed.get_entry(dataset.dataset_id).display_name, "Display only")
            self.assertEqual(dataset.name, "Original")
            output = Path(self.controller.metadata_default_directory()) / "template.xlsx"
            self.controller.create_metadata_excel_template(dataset, output)
            self.assertTrue(output.is_file())
            from openpyxl import load_workbook

            workbook = load_workbook(output, read_only=True)
            rows = tuple(workbook.active.iter_rows(values_only=True))
            workbook.close()
            self.assertEqual(rows[0], METADATA_TEMPLATE_HEADERS)
            self.assertEqual(tuple(row[0] for row in rows[1:]), ("sample-b", "sample-a"))
            removed = self.controller.remove_dataset(dataset.dataset_id)
            self.assertEqual(removed.dataset_count, 0)

    def test_project_metadata_template_adds_source_batch_for_duplicate_sample_ids(self) -> None:
        first = SequenceDataset(
            "first", "First", SourceType.AB1_TRIMMED,
            (SequenceRecord("R1", "ATGC", metadata={"source_batch": "Batch-A"}),),
        )
        second = SequenceDataset(
            "second", "Second", SourceType.AB1_TRIMMED,
            (SequenceRecord("R1", "ATGT", metadata={"source_batch": "Batch-B"}),),
        )
        with TemporaryDirectory() as directory:
            output = Path(directory) / "template.xlsx"
            write_project_metadata_excel_template((first, second), output)
            from openpyxl import load_workbook
            workbook = load_workbook(output, read_only=True)
            rows = tuple(workbook.active.iter_rows(values_only=True))
            workbook.close()
        self.assertEqual(rows[0][:2], ("Sample_ID", "Source_Batch"))
        self.assertEqual(rows[1][:2], ("R1", "Batch-A"))
        self.assertEqual(rows[2][:2], ("R1", "Batch-B"))

    def test_project_scope_metadata_revision_survives_save_reopen(self) -> None:
        first = SequenceDataset(
            "first", "First", SourceType.AB1_TRIMMED,
            (SequenceRecord("R1", "ATGC", metadata={"source_batch": "Batch-A"}),),
        )
        second = SequenceDataset(
            "second", "Second", SourceType.AB1_TRIMMED,
            (SequenceRecord("R1", "ATGT", metadata={"source_batch": "Batch-B"}),),
        )
        with TemporaryDirectory() as directory:
            project = self.controller.create_project("Metadata Scope", location=directory)
            self.state.replace_project(project.add_dataset(first).add_dataset(second))
            metadata = Path(directory) / "metadata.csv"
            metadata.write_text(
                "Sample_ID,Source_Batch,Location\nR1,Batch-A,Rembang\nR1,Batch-B,Cirebon\n",
                encoding="utf-8",
            )
            merged = self.controller.import_sample_metadata_for_project(str(metadata))
            expected = tuple(
                (dataset.dataset_id, dataset.records[0].sequence_id, dataset.records[0].metadata,
                 dataset.records[0].provenance)
                for dataset in merged
            )
            bundle = self.controller.save_project_bundle()
            reopened_state = AppState()
            reopened = ProjectController(reopened_state)
            reopened.open_project_bundle(bundle)
            actual = tuple(
                (entry.dataset.dataset_id, entry.dataset.records[0].sequence_id,
                 entry.dataset.records[0].metadata, entry.dataset.records[0].provenance)
                for entry in reopened_state.current_project.current_dataset_entries()
            )
            reopened_state.close_current_bundle()
        self.assertEqual(actual, expected)

    def test_remove_dataset_refuses_child_or_result_dependencies(self) -> None:
        root = _dataset("root", "Root")
        child = _dataset("child", "Child")
        project = Project.create("p", "Project").add_dataset(root).add_dataset(
            child, parent_dataset_id=root.dataset_id
        )
        self.state.set_project(project)
        with self.assertRaisesRegex(ValueError, "children exist"):
            self.controller.remove_dataset(root.dataset_id)
