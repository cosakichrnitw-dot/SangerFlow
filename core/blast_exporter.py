from openpyxl import Workbook
from openpyxl.styles import Font, Alignment

from core.blast_summary import (
    species_summary,
    get_best_hits
)



# ==================================================
# Extract accession
# ==================================================

def get_accession(title):

    import re


    match = re.search(
        r"\|([A-Z]{1,3}\d+\.\d+)\|",
        title
    )


    if match:

        return match.group(1)


    return "Unknown"





# ==================================================
# Export BLAST Excel
# ==================================================

def export_blast_excel(
    blast_results,
    quality_results,
    filepath
):


    wb = Workbook()



    # ==================================================
    # Sheet 1 : BLAST Result
    # ==================================================

    ws_blast = wb.active

    ws_blast.title = "BLAST_Result"



    blast_headers = [

        "Sample",
        "Species",
        "Identity (%)",
        "Coverage (%)",
        "Alignment Length",
        "E-value",
        "Accession",
        "Title"

    ]


    ws_blast.append(
        blast_headers
    )



    for cell in ws_blast[1]:

        cell.font = Font(
            bold=True
        )



    for result in blast_results:


        ws_blast.append(

            [

                result.get(
                    "sample",
                    ""
                ),

                result.get(
                    "species",
                    ""
                ),

                result.get(
                    "identity",
                    ""
                ),

                result.get(
                    "coverage",
                    ""
                ),

                result.get(
                    "alignment_length",
                    ""
                ),

                result.get(
                    "e_value",
                    ""
                ),

                get_accession(

                    result.get(
                        "title",
                        ""
                    )

                ),

                result.get(
                    "title",
                    ""
                )

            ]

        )





    # ==================================================
    # Sheet 2 : Quality Report
    # ==================================================

    ws_quality = wb.create_sheet(

        "Quality_Report"

    )



    quality_headers = [

        "Sample",
        "Original Length",
        "Average Quality",
        "Q20 (%)",
        "Q30 (%)",
        "HQ (%)",
        "Trim Start",
        "Trim End",
        "Trimmed Length"

    ]



    ws_quality.append(

        quality_headers

    )



    for cell in ws_quality[1]:

        cell.font = Font(
            bold=True
        )



    for q in quality_results:


        ws_quality.append(

            [

                q.get(
                    "filename",
                    ""
                ),

                q.get(
                    "length",
                    ""
                ),

                q.get(
                    "average_quality",
                    ""
                ),

                q.get(
                    "q20_rate",
                    ""
                ),

                q.get(
                    "q30_rate",
                    ""
                ),

                q.get(
                    "hq_percent",
                    ""
                ),

                q.get(
                    "trim_start",
                    ""
                ),

                q.get(
                    "trim_end",
                    ""
                ),

                q.get(
                    "trimmed_length",
                    ""
                )

            ]

        )





    # ==================================================
    # Sheet 3 : Species Summary
    # ==================================================

    ws_species = wb.create_sheet(

        "Species_Summary"

    )



    ws_species.append(

        [

            "Species",

            "Sample Count"

        ]

    )



    for cell in ws_species[1]:

        cell.font = Font(
            bold=True
        )



    summary = species_summary(

        blast_results

    )



    for item in summary:


        ws_species.append(

            [

                item["species"],

                item["count"]

            ]

        )





    # ==================================================
    # Sheet 4 : Best Identification
    # ==================================================

    ws_best = wb.create_sheet(

        "Best_Identification"

    )



    ws_best.append(

        [

            "Sample",

            "Species",

            "Identity (%)",

            "Coverage (%)",

            "Accession",

            "Title"

        ]

    )



    for cell in ws_best[1]:

        cell.font = Font(
            bold=True
        )



    best_hits = get_best_hits(

        blast_results

    )



    for result in best_hits:


        ws_best.append(

            [

                result.get(
                    "sample",
                    ""
                ),

                result.get(
                    "species",
                    ""
                ),

                result.get(
                    "identity",
                    ""
                ),

                result.get(
                    "coverage",
                    ""
                ),

                get_accession(

                    result.get(
                        "title",
                        ""
                    )

                ),

                result.get(
                    "title",
                    ""
                )

            ]

        )





    # ==================================================
    # Formatting
    # ==================================================

    for ws in wb:


        for column in ws.columns:


            max_length = max(

                len(
                    str(cell.value)
                )

                if cell.value

                else 0

                for cell in column

            )


            ws.column_dimensions[

                column[0].column_letter

            ].width = min(

                max_length + 3,

                50

            )



        for row in ws.iter_rows():


            for cell in row:

                cell.alignment = Alignment(

                    vertical="center"

                )



    wb.save(

        filepath

    )


    print(

        f"BLAST report saved: {filepath}"

    )