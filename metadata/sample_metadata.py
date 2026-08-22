"""Import CSV/XLSX sample metadata and immutably attach it to datasets."""

from __future__ import annotations

import csv
from dataclasses import dataclass
from datetime import date, datetime
from pathlib import Path
import re
from types import MappingProxyType
from typing import Mapping

from openpyxl import load_workbook

from core.sequence_dataset import SequenceDataset, SequenceRecord
from core.lineage import RecordProvenance, RecordRef


_SAMPLE_ID_COLUMN = "sample_id"
_SOURCE_BATCH_COLUMN = "source_batch"


def _freeze_metadata(value: Mapping[str, object] | None) -> Mapping[str, object]:
    if value is None:
        return MappingProxyType({})
    if not isinstance(value, Mapping):
        raise ValueError("metadata must be a mapping")
    return MappingProxyType(dict(value))


@dataclass(frozen=True)
class SampleMetadataRecord:
    """One immutable metadata row keyed by an existing SequenceRecord ID."""

    sample_id: str
    metadata: Mapping[str, object]

    def __post_init__(self) -> None:
        if not isinstance(self.sample_id, str) or not self.sample_id.strip():
            raise ValueError("sample_id must be a non-empty string")
        object.__setattr__(self, "metadata", _freeze_metadata(self.metadata))


@dataclass(frozen=True)
class SampleMetadataTable:
    """Ordered, immutable table imported from a CSV or XLSX metadata file."""

    records: tuple[SampleMetadataRecord, ...]
    source_filepath: str
    columns: tuple[str, ...]

    def __post_init__(self) -> None:
        records = tuple(self.records)
        if not records:
            raise ValueError("sample metadata table must contain at least one row")
        if any(not isinstance(record, SampleMetadataRecord) for record in records):
            raise ValueError("records must contain SampleMetadataRecord values")
        if not isinstance(self.source_filepath, str) or not self.source_filepath:
            raise ValueError("source_filepath must be a non-empty string")
        columns = tuple(self.columns)
        if _SAMPLE_ID_COLUMN not in columns:
            raise ValueError("metadata table requires a Sample_ID column")
        _validate_metadata_row_keys(records, columns)
        object.__setattr__(self, "records", records)
        object.__setattr__(self, "columns", columns)

    @property
    def sample_ids(self) -> tuple[str, ...]:
        return tuple(record.sample_id for record in self.records)

    def get_record(self, sample_id: str) -> SampleMetadataRecord:
        for record in self.records:
            if record.sample_id == sample_id:
                return record
        raise KeyError(sample_id)


def import_sample_metadata(filepath: str | Path) -> SampleMetadataTable:
    """Read a ``.csv`` or ``.xlsx`` metadata table requiring ``Sample_ID``."""
    path = _validate_input_path(filepath)
    suffix = path.suffix.lower()
    if suffix == ".csv":
        headers, rows = _read_csv(path)
    elif suffix == ".xlsx":
        headers, rows = _read_xlsx(path)
    else:
        raise ValueError("sample metadata filepath must use .csv or .xlsx")
    return _build_table(path, headers, rows)


def merge_sample_metadata(
    dataset: SequenceDataset,
    metadata_table: SampleMetadataTable,
) -> SequenceDataset:
    """Return a new dataset whose matching records contain imported metadata.

    Every metadata ``Sample_ID`` must match a SequenceRecord ID.  Dataset rows
    absent from the metadata file are retained unchanged, enabling partial
    annotation while still detecting accidental external sample IDs.
    """
    if not isinstance(dataset, SequenceDataset):
        raise ValueError("dataset must be a SequenceDataset")
    if not isinstance(metadata_table, SampleMetadataTable):
        raise ValueError("metadata_table must be a SampleMetadataTable")
    matched = _resolve_metadata_matches((dataset,), metadata_table)
    metadata_by_record = {record_key: metadata for record_key, metadata in matched.items()}
    records = tuple(
        SequenceRecord(
            sequence_id=record.sequence_id,
            sequence=record.sequence,
            description=record.description,
            source_reference=record.source_reference,
            metadata=_merge_record_metadata(
                record.metadata,
                metadata_by_record.get((dataset.dataset_id, record.sequence_id), {}),
            ),
            provenance=RecordProvenance(
                (RecordRef(dataset.dataset_id, record.sequence_id),)
            ),
        )
        for record in dataset.records
    )
    dataset_metadata = dict(dataset.metadata)
    dataset_metadata.update(
        {
            "sample_metadata_merged": True,
            "sample_metadata_source": metadata_table.source_filepath,
            "sample_metadata_matched_count": len(metadata_table.records),
        }
    )
    return SequenceDataset(
        dataset_id=dataset.dataset_id,
        name=dataset.name,
        source_type=dataset.source_type,
        records=records,
        metadata=dataset_metadata,
    )


def merge_sample_metadata_for_datasets(
    datasets: tuple[SequenceDataset, ...],
    metadata_table: SampleMetadataTable,
) -> tuple[SequenceDataset, ...]:
    """Merge one metadata file across an explicit multi-Dataset scope.

    This preserves each Dataset and record identity.  ``Source_Batch`` is only
    a matching discriminator and is never imported over a record's provenance
    metadata.
    """
    datasets = tuple(datasets)
    if not datasets or any(not isinstance(dataset, SequenceDataset) for dataset in datasets):
        raise ValueError("datasets must contain at least one SequenceDataset")
    matched = _resolve_metadata_matches(datasets, metadata_table)
    merged: list[SequenceDataset] = []
    for dataset in datasets:
        records = tuple(
            SequenceRecord(
                sequence_id=record.sequence_id,
                sequence=record.sequence,
                description=record.description,
                source_reference=record.source_reference,
                metadata=_merge_record_metadata(
                    record.metadata,
                    matched.get((dataset.dataset_id, record.sequence_id), {}),
                ),
                provenance=RecordProvenance((RecordRef(dataset.dataset_id, record.sequence_id),)),
            )
            for record in dataset.records
        )
        merged.append(SequenceDataset(
            dataset_id=dataset.dataset_id,
            name=dataset.name,
            source_type=dataset.source_type,
            records=records,
            metadata={
                **dict(dataset.metadata),
                "sample_metadata_merged": True,
                "sample_metadata_source": metadata_table.source_filepath,
                "sample_metadata_matched_count": sum(
                    1 for key in matched if key[0] == dataset.dataset_id
                ),
            },
        ))
    return tuple(merged)


def _merge_record_metadata(
    existing: Mapping[str, object],
    imported: Mapping[str, object],
) -> dict[str, object]:
    """Merge imported canonical keys without retaining case-only duplicates.

    Import headers are intentionally normalized (``Location`` -> ``location``).
    Replace an existing key that differs only by case/spacing so Project-wide
    discovery and filtering have one authoritative value per metadata field.
    """

    merged = dict(existing)
    imported = {
        key: value for key, value in imported.items()
        if _normalize_column_name(key) != _SOURCE_BATCH_COLUMN
    }
    imported_keys = {_normalize_column_name(key) for key in imported}
    for key in tuple(merged):
        if _normalize_column_name(key) in imported_keys:
            del merged[key]
    merged.update(imported)
    return merged


def _validate_metadata_row_keys(
    records: tuple[SampleMetadataRecord, ...],
    columns: tuple[str, ...],
) -> None:
    by_sample: dict[str, list[SampleMetadataRecord]] = {}
    for record in records:
        by_sample.setdefault(record.sample_id, []).append(record)
    duplicated = {sample_id: rows for sample_id, rows in by_sample.items() if len(rows) > 1}
    if not duplicated:
        return
    if _SOURCE_BATCH_COLUMN not in columns:
        raise ValueError("duplicate Sample_ID values are not allowed: " + ", ".join(sorted(duplicated)))
    duplicate_keys: list[str] = []
    missing_batches: list[str] = []
    for sample_id, rows in duplicated.items():
        batches = [str(row.metadata.get(_SOURCE_BATCH_COLUMN, "")).strip() for row in rows]
        if not all(batches):
            missing_batches.append(sample_id)
        elif len(set(batches)) != len(batches):
            duplicate_keys.append(sample_id)
    if missing_batches:
        raise ValueError("duplicate Sample_ID values require Source_Batch: " + ", ".join(sorted(missing_batches)))
    if duplicate_keys:
        raise ValueError("duplicate Sample_ID + Source_Batch values are not allowed: " + ", ".join(sorted(duplicate_keys)))


def _resolve_metadata_matches(
    datasets: tuple[SequenceDataset, ...],
    metadata_table: SampleMetadataTable,
) -> dict[tuple[str, str], Mapping[str, object]]:
    candidates: dict[str, list[tuple[SequenceDataset, SequenceRecord]]] = {}
    for dataset in datasets:
        for record in dataset.records:
            candidates.setdefault(record.sequence_id, []).append((dataset, record))
    matches: dict[tuple[str, str], Mapping[str, object]] = {}
    unmatched: list[str] = []
    ambiguous: list[str] = []
    for metadata_record in metadata_table.records:
        options = candidates.get(metadata_record.sample_id, [])
        if not options:
            unmatched.append(metadata_record.sample_id)
            continue
        if len(options) == 1:
            key = (options[0][0].dataset_id, options[0][1].sequence_id)
        else:
            if _SOURCE_BATCH_COLUMN not in metadata_table.columns:
                ambiguous.append(metadata_record.sample_id)
                continue
            batch = str(metadata_record.metadata.get(_SOURCE_BATCH_COLUMN, "")).strip()
            batch_options = [
                option for option in options
                if str(option[1].metadata.get(_SOURCE_BATCH_COLUMN, "")).strip() == batch
            ]
            if len(batch_options) != 1:
                (unmatched if not batch_options else ambiguous).append(metadata_record.sample_id)
                continue
            key = (batch_options[0][0].dataset_id, batch_options[0][1].sequence_id)
        if key in matches:
            ambiguous.append(metadata_record.sample_id)
            continue
        matches[key] = metadata_record.metadata
    if unmatched:
        raise ValueError("metadata contains unmatched Sample_ID/Source_Batch: " + ", ".join(sorted(set(unmatched))))
    if ambiguous:
        raise ValueError("ambiguous metadata match for Sample_ID: " + ", ".join(sorted(set(ambiguous))))
    return matches


def _validate_input_path(filepath: str | Path) -> Path:
    if not isinstance(filepath, (str, Path)):
        raise ValueError("filepath must be a path string or Path")
    path = Path(filepath)
    if not path.is_file():
        raise ValueError(f"sample metadata file does not exist: {path}")
    return path


def _read_csv(path: Path) -> tuple[tuple[str, ...], tuple[tuple[object, ...], ...]]:
    try:
        with path.open("r", encoding="utf-8-sig", newline="") as handle:
            reader = csv.reader(handle)
            rows = tuple(tuple(row) for row in reader)
    except OSError as error:
        raise ValueError(f"could not read sample metadata CSV: {error}") from error
    if not rows:
        raise ValueError("sample metadata CSV is empty")
    return tuple(rows[0]), tuple(row for row in rows[1:] if any(str(value).strip() for value in row))


def _read_xlsx(path: Path) -> tuple[tuple[str, ...], tuple[tuple[object, ...], ...]]:
    try:
        workbook = load_workbook(path, read_only=True, data_only=True)
        worksheet = workbook.active
        rows = tuple(tuple(row) for row in worksheet.iter_rows(values_only=True))
        workbook.close()
    except OSError as error:
        raise ValueError(f"could not read sample metadata XLSX: {error}") from error
    if not rows:
        raise ValueError("sample metadata XLSX is empty")
    headers = tuple("" if value is None else str(value) for value in rows[0])
    data_rows = tuple(row for row in rows[1:] if any(value is not None and str(value).strip() for value in row))
    return headers, data_rows


def _build_table(
    path: Path,
    headers: tuple[str, ...],
    rows: tuple[tuple[object, ...], ...],
) -> SampleMetadataTable:
    normalized_headers = tuple(_normalize_column_name(header) for header in headers)
    if not normalized_headers or any(not header for header in normalized_headers):
        raise ValueError("metadata headers must be non-empty")
    if len(set(normalized_headers)) != len(normalized_headers):
        raise ValueError("duplicate metadata columns are not allowed")
    if _SAMPLE_ID_COLUMN not in normalized_headers:
        raise ValueError("metadata requires a Sample_ID column")
    if not rows:
        raise ValueError("sample metadata file contains no rows")
    sample_id_index = normalized_headers.index(_SAMPLE_ID_COLUMN)
    records = []
    for row_number, row in enumerate(rows, start=2):
        padded_row = tuple(row) + (None,) * max(0, len(normalized_headers) - len(row))
        if len(row) > len(normalized_headers):
            raise ValueError(f"metadata row {row_number} has more values than headers")
        sample_id_value = padded_row[sample_id_index]
        sample_id = "" if sample_id_value is None else str(sample_id_value).strip()
        if not sample_id:
            raise ValueError(f"metadata row {row_number} has an empty Sample_ID")
        metadata = {
            column: _normalize_cell_value(value, column)
            for column, value in zip(normalized_headers, padded_row)
            if column != _SAMPLE_ID_COLUMN and value is not None and str(value).strip() != ""
        }
        records.append(SampleMetadataRecord(sample_id=sample_id, metadata=metadata))
    return SampleMetadataTable(
        records=tuple(records),
        source_filepath=str(path),
        columns=normalized_headers,
    )


def _normalize_column_name(value: object) -> str:
    if not isinstance(value, str):
        return ""
    normalized = re.sub(r"[^a-z0-9]+", "_", value.strip().lower()).strip("_")
    return normalized


def _normalize_cell_value(value: object, column: str) -> object:
    if isinstance(value, datetime):
        return value.date().isoformat() if column == "collection_date" else value.isoformat()
    if isinstance(value, date):
        return value.isoformat()
    if column in {"latitude", "longitude"} and isinstance(value, str):
        try:
            return float(value)
        except ValueError as error:
            raise ValueError(f"{column} must be numeric when supplied") from error
    return value
