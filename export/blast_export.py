"""Excel and TSV export adapters for immutable BLAST result datasets."""

from __future__ import annotations

from collections.abc import Iterator
import csv
from pathlib import Path
from typing import Literal

from openpyxl import Workbook
from openpyxl.styles import Font

from core.blast_result import BlastHit, BlastResultDataset


_ALL_HIT_HEADERS = (
    "query_id",
    "rank",
    "hit_accession",
    "scientific_name",
    "organism",
    "identity",
    "query_coverage",
    "evalue",
    "alignment_length",
    "database",
    "bit_score",
    "description",
)

BLAST_EXPORT_COLUMNS = (
    "query_id", "rank", "scientific_name", "organism", "description",
    "hit_accession", "identity", "query_coverage", "alignment_length",
    "evalue", "bit_score", "database",
)
IDENTIFICATION_SUMMARY_COLUMNS = (
    "query_id", "rank", "scientific_name", "hit_accession", "identity",
    "query_coverage", "evalue", "bit_score",
)

_QUERY_SUMMARY_HEADERS = (
    "query_id",
    "top_accession",
    "top_scientific_name",
    "top_organism",
    "top_identity",
    "top_coverage",
    "top_evalue",
    "top_alignment_length",
)


def export_blast_result_to_excel(
    blast_result: BlastResultDataset,
    filepath: str | Path,
) -> None:
    """Write one immutable BLAST result dataset to a three-sheet workbook."""

    _validate_blast_result(blast_result)
    output_path = _validate_output_path(filepath, suffix=".xlsx")

    workbook = Workbook()
    summary_sheet = workbook.active
    summary_sheet.title = "Summary"
    _write_summary(summary_sheet, blast_result)

    query_summary_sheet = workbook.create_sheet("Query Summary")
    _write_query_summary(query_summary_sheet, blast_result)

    all_hits_sheet = workbook.create_sheet("All Hits")
    _write_all_hits(all_hits_sheet, blast_result)

    workbook.save(output_path)


def export_blast_result_to_tsv(
    blast_result: BlastResultDataset,
    filepath: str | Path,
) -> None:
    """Write all BLAST hits in query-local rank order as tab-separated data."""

    _validate_blast_result(blast_result)
    output_path = _validate_output_path(filepath, suffix=".tsv")
    with output_path.open("w", encoding="utf-8", newline="") as handle:
        handle.write("\t".join(_ALL_HIT_HEADERS) + "\n")
        for query_id, rank, hit in _ranked_hits(blast_result):
            values = _all_hit_values(query_id, rank, hit)
            handle.write("\t".join(str(value) for value in values) + "\n")


def export_blast_result_to_csv(
    blast_result: BlastResultDataset,
    filepath: str | Path,
    *,
    columns: tuple[str, ...] = BLAST_EXPORT_COLUMNS,
) -> None:
    """Write one UTF-8 CSV row per ranked hit using selected supported columns."""

    _validate_blast_result(blast_result)
    selected = _validate_columns(columns)
    output_path = _validate_output_path(filepath, suffix=".csv")
    with output_path.open("w", encoding="utf-8", newline="") as handle:
        writer = csv.writer(handle)
        writer.writerow(selected)
        for query_id, rank, hit in _ranked_hits(blast_result):
            values = _hit_mapping(query_id, rank, hit)
            writer.writerow([values[column] for column in selected])


def export_blast_result_to_excel_selected(
    blast_result: BlastResultDataset,
    filepath: str | Path,
    *,
    columns: tuple[str, ...] = BLAST_EXPORT_COLUMNS,
) -> None:
    """Write the compatible Summary/Query Summary sheets plus selected All Hits."""

    _validate_blast_result(blast_result)
    selected = _validate_columns(columns)
    output_path = _validate_output_path(filepath, suffix=".xlsx")
    workbook = Workbook()
    summary_sheet = workbook.active
    summary_sheet.title = "Summary"
    _write_summary(summary_sheet, blast_result)
    query_summary_sheet = workbook.create_sheet("Query Summary")
    _write_query_summary(query_summary_sheet, blast_result)
    hits_sheet = workbook.create_sheet("All Hits")
    hits_sheet.append(selected)
    for query_id, rank, hit in _ranked_hits(blast_result):
        values = _hit_mapping(query_id, rank, hit)
        hits_sheet.append(tuple(values[column] for column in selected))
    _bold_header_rows(hits_sheet, (1,))
    workbook.save(output_path)


def _validate_blast_result(blast_result: object) -> BlastResultDataset:
    if not isinstance(blast_result, BlastResultDataset):
        raise ValueError("blast_result must be a BlastResultDataset")
    if not blast_result.hits:
        raise ValueError("blast_result must contain at least one hit to export")
    return blast_result


def _validate_output_path(
    filepath: str | Path,
    *,
    suffix: Literal[".xlsx", ".tsv", ".csv"],
) -> Path:
    if not isinstance(filepath, (str, Path)):
        raise ValueError("filepath must be a path string or Path")
    output_path = Path(filepath)
    if not output_path.name or output_path.suffix.lower() != suffix:
        raise ValueError(f"filepath must use the {suffix} extension")
    if output_path.exists() and output_path.is_dir():
        raise ValueError("filepath must refer to a file, not a directory")
    if not output_path.parent.exists() or not output_path.parent.is_dir():
        raise ValueError("filepath parent directory does not exist")
    return output_path


def _write_summary(sheet: object, blast_result: BlastResultDataset) -> None:
    # ``sheet`` is deliberately kept duck-typed to avoid exposing openpyxl
    # worksheet details in this result-model adapter's public surface.
    rows = (
        ("result_id", blast_result.result_id),
        ("result_name", blast_result.name),
        ("analysis_mode", blast_result.analysis_mode.value),
        ("marker", blast_result.marker or ""),
        ("database", blast_result.database or ""),
        ("parent_dataset_id", blast_result.parent_dataset_id),
        ("query_count", len(blast_result.query_ids())),
        ("hit_count", blast_result.hit_count()),
    )
    sheet.append(("field", "value"))  # type: ignore[attr-defined]
    for row in rows:
        sheet.append(row)  # type: ignore[attr-defined]
    sheet.append(())  # type: ignore[attr-defined]
    sheet.append(("result_metadata_key", "result_metadata_value"))  # type: ignore[attr-defined]
    for key, value in sorted(blast_result.metadata.items()):
        sheet.append((key, str(value)))  # type: ignore[attr-defined]
    _bold_header_rows(sheet, (1, 11))


def _write_query_summary(sheet: object, blast_result: BlastResultDataset) -> None:
    sheet.append(_QUERY_SUMMARY_HEADERS)  # type: ignore[attr-defined]
    for query_id in blast_result.query_ids():
        top_hit = blast_result.get_hits(query_id)[0]
        sheet.append(  # type: ignore[attr-defined]
            (
                query_id,
                top_hit.hit_accession,
                top_hit.scientific_name,
                top_hit.organism,
                top_hit.identity,
                top_hit.query_coverage,
                top_hit.evalue,
                top_hit.alignment_length,
            )
        )
    _bold_header_rows(sheet, (1,))


def _write_all_hits(sheet: object, blast_result: BlastResultDataset) -> None:
    sheet.append(_ALL_HIT_HEADERS)  # type: ignore[attr-defined]
    for query_id, rank, hit in _ranked_hits(blast_result):
        sheet.append(_all_hit_values(query_id, rank, hit))  # type: ignore[attr-defined]
    _bold_header_rows(sheet, (1,))


def _bold_header_rows(sheet: object, rows: tuple[int, ...]) -> None:
    for row_index in rows:
        for cell in sheet[row_index]:  # type: ignore[index]
            cell.font = Font(bold=True)


def _ranked_hits(blast_result: BlastResultDataset) -> Iterator[tuple[str, int, BlastHit]]:
    ranks: dict[str, int] = {}
    for hit in blast_result.hits:
        rank = ranks.get(hit.query_id, 0) + 1
        ranks[hit.query_id] = rank
        yield hit.query_id, rank, hit


def _all_hit_values(query_id: str, rank: int, hit: BlastHit) -> tuple[object, ...]:
    return (
        query_id,
        rank,
        hit.hit_accession,
        hit.scientific_name,
        hit.organism,
        hit.identity,
        hit.query_coverage,
        hit.evalue,
        hit.alignment_length,
        hit.database,
        hit.bit_score if hit.bit_score is not None else "",
        hit.description or "",
    )


def _hit_mapping(query_id: str, rank: int, hit: BlastHit) -> dict[str, object]:
    return dict(zip(_ALL_HIT_HEADERS, _all_hit_values(query_id, rank, hit), strict=True))


def _validate_columns(columns: tuple[str, ...]) -> tuple[str, ...]:
    selected = tuple(columns)
    if not selected or len(set(selected)) != len(selected) or any(column not in BLAST_EXPORT_COLUMNS for column in selected):
        raise ValueError("columns must be a non-empty unique subset of supported BLAST export columns")
    return selected
