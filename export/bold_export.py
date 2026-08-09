"""Excel and TSV export adapters for immutable BOLD result datasets."""

from __future__ import annotations

from pathlib import Path
from typing import Literal

from openpyxl import Workbook
from openpyxl.styles import Font

from core.bold_result import BoldHit, BoldResultDataset


_TAXONOMIC_ASSIGNMENT_HEADERS = (
    "query_id",
    "species_name",
    "genus",
    "family",
    "order",
    "phylum",
    "similarity",
)

_REFERENCE_INFORMATION_HEADERS = (
    "query_id",
    "process_id",
    "record_id",
    "bin_uri",
    "country",
    "institution",
    "specimen_id",
    "collection_date",
)

_ALL_HITS_HEADERS = (
    "query_id",
    "process_id",
    "record_id",
    "species_name",
    "similarity",
    "bin_uri",
    "country",
    "database",
)


def export_bold_result_to_excel(
    bold_result: BoldResultDataset,
    filepath: str | Path,
) -> None:
    """Write a BOLD result dataset to a four-sheet ``.xlsx`` workbook."""

    _validate_bold_result(bold_result)
    output_path = _validate_output_path(filepath, suffix=".xlsx")

    workbook = Workbook()
    summary_sheet = workbook.active
    summary_sheet.title = "Summary"
    _write_summary(summary_sheet, bold_result)

    taxonomic_sheet = workbook.create_sheet("Taxonomic Assignment")
    _write_taxonomic_assignments(taxonomic_sheet, bold_result)

    reference_sheet = workbook.create_sheet("Reference Information")
    _write_reference_information(reference_sheet, bold_result)

    all_hits_sheet = workbook.create_sheet("All Hits")
    _write_all_hits(all_hits_sheet, bold_result)

    try:
        workbook.save(output_path)
    except OSError as error:
        raise ValueError(f"unable to write Excel file: {output_path}") from error


def export_bold_result_to_tsv(
    bold_result: BoldResultDataset,
    filepath: str | Path,
) -> None:
    """Write BOLD hits as an ``All Hits``-format tab-separated file."""

    _validate_bold_result(bold_result)
    output_path = _validate_output_path(filepath, suffix=".tsv")
    try:
        with output_path.open("w", encoding="utf-8", newline="") as handle:
            handle.write("\t".join(_ALL_HITS_HEADERS) + "\n")
            for hit in bold_result.hits:
                handle.write("\t".join(_stringify(value) for value in _all_hit_values(hit)) + "\n")
    except OSError as error:
        raise ValueError(f"unable to write TSV file: {output_path}") from error


def _validate_bold_result(bold_result: object) -> BoldResultDataset:
    if not isinstance(bold_result, BoldResultDataset):
        raise ValueError("bold_result must be a BoldResultDataset")
    if not bold_result.hits:
        raise ValueError("bold_result must contain at least one hit to export")
    return bold_result


def _validate_output_path(
    filepath: str | Path,
    *,
    suffix: Literal[".xlsx", ".tsv"],
) -> Path:
    if not isinstance(filepath, (str, Path)):
        raise ValueError("filepath must be a path string or Path")
    output_path = Path(filepath)
    if not output_path.name or output_path.suffix.lower() != suffix:
        raise ValueError(f"filepath must use the {suffix} extension")
    if output_path.exists() and output_path.is_dir():
        raise ValueError("filepath must refer to a file, not a directory")
    if not output_path.parent.exists() or not output_path.parent.is_dir():
        raise ValueError("filepath parent directory does not exist")
    return output_path


def _write_summary(sheet: object, bold_result: BoldResultDataset) -> None:
    rows = (
        ("result_id", bold_result.result_id),
        ("name", bold_result.name),
        ("parent_dataset_id", bold_result.parent_dataset_id),
        ("marker", bold_result.marker or ""),
        ("database", bold_result.database),
        ("query_count", len(bold_result.query_ids())),
        ("hit_count", bold_result.hit_count()),
    )
    sheet.append(("field", "value"))  # type: ignore[attr-defined]
    for row in rows:
        sheet.append(row)  # type: ignore[attr-defined]
    sheet.append(())  # type: ignore[attr-defined]
    sheet.append(("result_metadata_key", "result_metadata_value"))  # type: ignore[attr-defined]
    for key, value in sorted(bold_result.metadata.items()):
        sheet.append((key, _stringify(value)))  # type: ignore[attr-defined]
    _bold_header_rows(sheet, (1, 10))


def _write_taxonomic_assignments(sheet: object, bold_result: BoldResultDataset) -> None:
    sheet.append(_TAXONOMIC_ASSIGNMENT_HEADERS)  # type: ignore[attr-defined]
    for hit in bold_result.hits:
        sheet.append(  # type: ignore[attr-defined]
            (
                hit.query_id,
                hit.species_name,
                hit.genus,
                hit.family,
                hit.order,
                hit.phylum,
                hit.similarity,
            )
        )
    _bold_header_rows(sheet, (1,))


def _write_reference_information(sheet: object, bold_result: BoldResultDataset) -> None:
    sheet.append(_REFERENCE_INFORMATION_HEADERS)  # type: ignore[attr-defined]
    for hit in bold_result.hits:
        sheet.append(  # type: ignore[attr-defined]
            (
                hit.query_id,
                hit.process_id,
                hit.record_id,
                hit.bin_uri,
                hit.country,
                hit.institution,
                hit.specimen_id,
                hit.collection_date,
            )
        )
    _bold_header_rows(sheet, (1,))


def _write_all_hits(sheet: object, bold_result: BoldResultDataset) -> None:
    sheet.append(_ALL_HITS_HEADERS)  # type: ignore[attr-defined]
    for hit in bold_result.hits:
        sheet.append(_all_hit_values(hit))  # type: ignore[attr-defined]
    _bold_header_rows(sheet, (1,))


def _all_hit_values(hit: BoldHit) -> tuple[object, ...]:
    return (
        hit.query_id,
        hit.process_id,
        hit.record_id,
        hit.species_name,
        hit.similarity,
        hit.bin_uri,
        hit.country,
        hit.database,
    )


def _stringify(value: object) -> str:
    return "" if value is None else str(value)


def _bold_header_rows(sheet: object, rows: tuple[int, ...]) -> None:
    for row_index in rows:
        for cell in sheet[row_index]:  # type: ignore[index]
            cell.font = Font(bold=True)
