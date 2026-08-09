"""Tests for BOLD-result Excel and TSV export adapters."""

from __future__ import annotations

from dataclasses import FrozenInstanceError
from pathlib import Path
from tempfile import TemporaryDirectory
import unittest

from openpyxl import load_workbook

from core.bold_result import BoldHit, BoldResultDataset
from export.bold_export import export_bold_result_to_excel, export_bold_result_to_tsv


def make_result() -> BoldResultDataset:
    return BoldResultDataset(
        result_id="coi-bold",
        name="COI BOLD identification",
        parent_dataset_id="coi-trimmed",
        marker="COI",
        database="BOLD",
        hits=(
            BoldHit(
                query_id="IK345",
                process_id="BOLD:AAA001",
                record_id="REC-001",
                species_name="Rhynchobatus australiae",
                genus="Rhynchobatus",
                family="Rhinidae",
                order="Rhinopristiformes",
                phylum="Chordata",
                bin_uri="BOLD:AAA001",
                similarity=99.4,
                database="BOLD",
                country="Indonesia",
                institution="Museum A",
                specimen_id="SPEC-001",
                collection_date="2026-08-05",
            ),
            BoldHit(
                query_id="IK346",
                process_id="BOLD:BBB002",
                record_id="REC-002",
                species_name="Rhynchobatus palpebratus",
                genus="Rhynchobatus",
                family="Rhinidae",
                order="Rhinopristiformes",
                phylum="Chordata",
                bin_uri="BOLD:BBB002",
                similarity=98.8,
                database="BOLD",
                country="Malaysia",
                institution="Museum B",
                specimen_id="SPEC-002",
                collection_date="2025-01-15",
            ),
        ),
        metadata={"workflow": "BOLD", "run_label": "validation"},
    )


class BoldExportTests(unittest.TestCase):
    def test_excel_contains_all_required_sheets_and_bold_information(self) -> None:
        result = make_result()
        with TemporaryDirectory() as directory:
            path = Path(directory) / "bold.xlsx"
            export_bold_result_to_excel(result, path)

            self.assertTrue(path.is_file())
            workbook = load_workbook(path, data_only=True)
            self.assertEqual(
                workbook.sheetnames,
                ["Summary", "Taxonomic Assignment", "Reference Information", "All Hits"],
            )
            summary = workbook["Summary"]
            summary_values = {
                summary.cell(row=row, column=1).value: summary.cell(row=row, column=2).value
                for row in range(2, 9)
            }
            self.assertEqual(summary_values["result_id"], "coi-bold")
            self.assertEqual(summary_values["name"], "COI BOLD identification")
            self.assertEqual(summary_values["parent_dataset_id"], "coi-trimmed")
            self.assertEqual(summary_values["marker"], "COI")
            self.assertEqual(summary_values["database"], "BOLD")
            self.assertEqual(summary_values["query_count"], 2)
            self.assertEqual(summary_values["hit_count"], 2)
            self.assertEqual(summary.cell(row=11, column=1).value, "run_label")
            self.assertEqual(summary.cell(row=11, column=2).value, "validation")

            taxonomy = workbook["Taxonomic Assignment"]
            self.assertEqual(taxonomy.cell(row=2, column=1).value, "IK345")
            self.assertEqual(taxonomy.cell(row=2, column=2).value, "Rhynchobatus australiae")
            self.assertEqual(taxonomy.cell(row=2, column=7).value, 99.4)

            references = workbook["Reference Information"]
            self.assertEqual(references.cell(row=2, column=2).value, "BOLD:AAA001")
            self.assertEqual(references.cell(row=2, column=4).value, "BOLD:AAA001")
            self.assertEqual(references.cell(row=2, column=6).value, "Museum A")
            self.assertEqual(references.cell(row=2, column=8).value, "2026-08-05")

            all_hits = workbook["All Hits"]
            self.assertEqual(all_hits.max_row, 3)
            self.assertEqual(all_hits.cell(row=3, column=1).value, "IK346")
            self.assertEqual(all_hits.cell(row=3, column=8).value, "BOLD")

    def test_tsv_uses_all_hits_structure(self) -> None:
        result = make_result()
        with TemporaryDirectory() as directory:
            path = Path(directory) / "bold.tsv"
            export_bold_result_to_tsv(result, path)

            lines = path.read_text(encoding="utf-8").splitlines()
            self.assertEqual(
                lines[0].split("\t"),
                [
                    "query_id", "process_id", "record_id", "species_name", "similarity",
                    "bin_uri", "country", "database",
                ],
            )
            self.assertEqual(
                lines[1].split("\t"),
                [
                    "IK345", "BOLD:AAA001", "REC-001", "Rhynchobatus australiae",
                    "99.4", "BOLD:AAA001", "Indonesia", "BOLD",
                ],
            )
            self.assertEqual(lines[2].split("\t")[0], "IK346")

    def test_rejects_invalid_input_empty_results_and_paths_without_mutating_result(self) -> None:
        result = make_result()
        empty = BoldResultDataset("empty", "Empty", "coi-trimmed", "COI", "BOLD", ())
        with TemporaryDirectory() as directory:
            with self.assertRaisesRegex(ValueError, "BoldResultDataset"):
                export_bold_result_to_excel(object(), Path(directory) / "bold.xlsx")  # type: ignore[arg-type]
            with self.assertRaisesRegex(ValueError, "at least one hit"):
                export_bold_result_to_tsv(empty, Path(directory) / "empty.tsv")
            with self.assertRaisesRegex(ValueError, "extension"):
                export_bold_result_to_excel(result, Path(directory) / "bold.tsv")
            with self.assertRaisesRegex(ValueError, "parent directory"):
                export_bold_result_to_tsv(result, Path(directory) / "missing" / "bold.tsv")
        self.assertEqual(result.hit_count(), 2)
        self.assertEqual(result.hits[0].bin_uri, "BOLD:AAA001")
        with self.assertRaises(FrozenInstanceError):
            result.name = "Changed"  # type: ignore[misc]


if __name__ == "__main__":
    unittest.main()
