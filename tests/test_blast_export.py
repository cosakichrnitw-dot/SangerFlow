"""Tests for BLAST-result Excel and TSV export adapters."""

from __future__ import annotations

from dataclasses import FrozenInstanceError
from pathlib import Path
from tempfile import TemporaryDirectory
import unittest

from openpyxl import load_workbook

from core.blast_result import BlastAnalysisMode, BlastHit, BlastResultDataset
from export.blast_export import export_blast_result_to_excel, export_blast_result_to_tsv


def make_hit(query_id: str, accession: str, identity: float) -> BlastHit:
    return BlastHit(
        query_id=query_id,
        hit_accession=accession,
        scientific_name="Rhynchobatus australiae",
        organism="Rhynchobatus australiae",
        identity=identity,
        query_coverage=98.0,
        evalue=1e-50,
        alignment_length=658,
        database="nt",
    )


def make_result() -> BlastResultDataset:
    return BlastResultDataset(
        result_id="coi-identification",
        name="COI identification BLAST",
        hits=(
            make_hit("IK345", "AB-first", 99.5),
            make_hit("IK345", "AB-second", 98.5),
            make_hit("IK346", "CD-first", 97.5),
        ),
        parent_dataset_id="coi-trimmed",
        analysis_mode=BlastAnalysisMode.IDENTIFICATION,
        marker="COI",
        database="nt",
        metadata={"workflow": "BLAST", "run_label": "validation"},
    )


class BlastExportTests(unittest.TestCase):
    def test_excel_contains_summary_query_summary_and_all_hits(self) -> None:
        result = make_result()
        with TemporaryDirectory() as directory:
            path = Path(directory) / "blast.xlsx"
            export_blast_result_to_excel(result, path)

            self.assertTrue(path.is_file())
            workbook = load_workbook(path, data_only=True)
            self.assertEqual(workbook.sheetnames, ["Summary", "Query Summary", "All Hits"])
            summary = workbook["Summary"]
            summary_values = {
                summary.cell(row=row, column=1).value: summary.cell(row=row, column=2).value
                for row in range(2, 10)
            }
            self.assertEqual(summary_values["result_id"], "coi-identification")
            self.assertEqual(summary_values["analysis_mode"], "IDENTIFICATION")
            self.assertEqual(summary_values["marker"], "COI")
            self.assertEqual(summary_values["database"], "nt")
            self.assertEqual(summary_values["parent_dataset_id"], "coi-trimmed")
            self.assertEqual(summary_values["query_count"], 2)
            self.assertEqual(summary_values["hit_count"], 3)

            query_summary = workbook["Query Summary"]
            self.assertEqual(query_summary.max_row, 3)
            self.assertEqual(query_summary.cell(row=2, column=1).value, "IK345")
            self.assertEqual(query_summary.cell(row=2, column=2).value, "AB-first")
            self.assertEqual(query_summary.cell(row=3, column=1).value, "IK346")

            all_hits = workbook["All Hits"]
            self.assertEqual(all_hits.max_row, 4)
            self.assertEqual(all_hits.cell(row=2, column=1).value, "IK345")
            self.assertEqual(all_hits.cell(row=2, column=2).value, 1)
            self.assertEqual(all_hits.cell(row=3, column=2).value, 2)
            self.assertEqual(all_hits.cell(row=4, column=1).value, "IK346")
            self.assertEqual(all_hits.cell(row=4, column=2).value, 1)

    def test_tsv_contains_all_hits_with_query_local_ranks(self) -> None:
        result = make_result()
        with TemporaryDirectory() as directory:
            path = Path(directory) / "blast.tsv"
            export_blast_result_to_tsv(result, path)

            lines = path.read_text(encoding="utf-8").splitlines()
            self.assertEqual(lines[0].split("\t"), [
                "query_id", "rank", "hit_accession", "scientific_name", "organism",
                "identity", "query_coverage", "evalue", "alignment_length", "database",
            ])
            self.assertEqual(lines[1].split("\t")[:3], ["IK345", "1", "AB-first"])
            self.assertEqual(lines[2].split("\t")[:3], ["IK345", "2", "AB-second"])
            self.assertEqual(lines[3].split("\t")[:3], ["IK346", "1", "CD-first"])

    def test_rejects_invalid_input_empty_hits_and_invalid_paths_without_mutating_result(self) -> None:
        result = make_result()
        empty_result = BlastResultDataset(
            "empty", "Empty", (), "input", analysis_mode=BlastAnalysisMode.QC
        )
        with TemporaryDirectory() as directory:
            with self.assertRaisesRegex(ValueError, "BlastResultDataset"):
                export_blast_result_to_excel(object(), Path(directory) / "result.xlsx")  # type: ignore[arg-type]
            with self.assertRaisesRegex(ValueError, "at least one hit"):
                export_blast_result_to_tsv(empty_result, Path(directory) / "empty.tsv")
            with self.assertRaisesRegex(ValueError, "extension"):
                export_blast_result_to_excel(result, Path(directory) / "result.tsv")
            with self.assertRaisesRegex(ValueError, "parent directory"):
                export_blast_result_to_tsv(result, Path(directory) / "missing" / "result.tsv")
        self.assertEqual(result.hit_count(), 3)
        self.assertEqual(result.hits[0].hit_accession, "AB-first")
        with self.assertRaises(FrozenInstanceError):
            result.name = "Changed"  # type: ignore[misc]


if __name__ == "__main__":
    unittest.main()
