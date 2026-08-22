from __future__ import annotations

from pathlib import Path
from tempfile import TemporaryDirectory
import unittest

from openpyxl import load_workbook

from core.sequence_dataset import SequenceDataset, SequenceRecord, SourceType
from export.metadata_export import export_dataset_metadata_to_csv, export_dataset_metadata_to_xlsx


class MetadataExportTests(unittest.TestCase):
    def test_csv_and_xlsx_have_standard_and_selected_dynamic_fields(self) -> None:
        dataset = SequenceDataset("d", "Dataset", SourceType.REVIEWED_CONSENSUS, (
            SequenceRecord("sample1", "ATGC", metadata={"location": "Cirebon", "voucher": "V-1"}),
        ))
        with TemporaryDirectory() as directory:
            csv_path = Path(directory) / "metadata.csv"
            xlsx_path = Path(directory) / "metadata.xlsx"
            export_dataset_metadata_to_csv(dataset, csv_path, fields=("location",))
            export_dataset_metadata_to_xlsx(dataset, xlsx_path, fields=("voucher",))
            self.assertEqual(csv_path.read_text(encoding="utf-8").splitlines()[0], "Sample_ID,Sequence_length,Source_type,location")
            sheet = load_workbook(xlsx_path, data_only=True)["Metadata"]
            self.assertEqual([cell.value for cell in sheet[1]], ["Sample_ID", "Sequence_length", "Source_type", "voucher"])
            self.assertEqual(sheet.cell(2, 4).value, "V-1")
